#!/usr/bin/env python3
"""The book as a workbook you can sort, pivot and argue with offline.

Not a backup — the ledger is already the audit trail, and it is the only copy that
matters. This exists to answer questions the daily Telegram message structurally cannot,
because they are questions ABOUT the history rather than about today.

The sheet that earns the file is **BoardFires**: every momentum-board candidate the panel
can reconstruct, with its forward excess return, and a `taken` flag joined from the
ledger. That makes the central question answerable for the first time — *did taking four
of the eight names the board offered beat taking all eight?* Selection is the one part of
this system that has never been measured, because the board's edge was validated on ALL
its candidates and the book only ever held a few of them. The same rows also carry sector
and holding-period behaviour across ~1,500 IDX events, which is the market-learning half.

BoardFires is built from `trade_backtest.build_candidates()` — the same deterministic
reconstruction the backtest uses — so the sheet cannot drift from the validated
definition. If the row count ever stops matching, one of them is wrong and it matters.

Contains lots, rupiah and equity, so it is subject to exactly the `.gitignore`
prohibition the portfolio page is: never inside the public repo tree. `--out` refuses a
path under it.

Usage:
    py scripts/export_book.py                       # -> data/book/exports/book-<date>.xlsx
    py scripts/export_book.py --out /tmp/book.xlsx
    py scripts/export_book.py --no-boardfires       # fast; skips the panel reconstruction
"""
from __future__ import annotations

import argparse
import json
import statistics
import sys
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

import position_book as pb  # noqa: E402
from alpha_lib import Panel  # noqa: E402
from live_prices import WIB  # noqa: E402

ROOT = Path(__file__).resolve().parents[1]
PUBLIC_DOCS = ROOT / "docs"

HDR = {"font": True, "fill": "FF1F2937"}


def _refuse_public(out: Path) -> None:
    """The workbook must never be written where a git add could reach it.

    Mechanical, not aspirational: .gitignore already forbids data/book/ in the public
    repo, but a file written into docs/ would be published by the daily run's own
    path-scoped commit without anyone noticing.
    """
    try:
        out.resolve().relative_to(PUBLIC_DOCS.resolve())
    except ValueError:
        return
    raise SystemExit(f"[!!] refusing to write the book into the PUBLIC docs/ tree: {out}\n"
                     f"     It carries lots, rupiah and equity. Pick a path outside it.")


# ------------------------------------------------------------------------------ sheets

def sheet_trades(state: dict) -> tuple:
    cols = ["symbol", "sector", "opened", "closed", "lots_initial", "entry_px", "exit_px",
            "r_idr", "realised_R", "realised_idr", "fees_idr", "held_days",
            "exit_reason", "entry_rule", "stop_basis"]
    rows = []
    for c in state["closed"]:
        held = ""
        try:
            held = (datetime.fromisoformat(c["closed"]).date()
                    - datetime.fromisoformat(c["opened"]).date()).days
        except (ValueError, KeyError, TypeError):
            pass
        rows.append([c.get("symbol"), c.get("sector"), c.get("opened"), c.get("closed"),
                     c.get("lots_initial"), c.get("entry_px"), c.get("exit_px"),
                     c.get("r_idr"), c.get("realised_r_total"), c.get("realised_idr"),
                     c.get("fees_idr"), held, c.get("exit_reason"),
                     c.get("entry_rule"), c.get("stop_basis")])
    return cols, rows


def sheet_open(state: dict, p: Panel, marks: dict) -> tuple:
    cols = ["symbol", "sector", "opened", "lots", "entry_px", "stop_px", "mark",
            "R", "unrealised_idr", "notional_idr", "r_idr", "stop_gap_pct",
            "entry_rule", "mark_source"]
    rows = []
    i = len(p.dates) - 1 if p.dates else None
    for pos in sorted(state["positions"].values(), key=lambda x: x["symbol"]):
        sym = pos["symbol"]
        q = marks.get(sym)
        mark = (q or {}).get("px") or (p.raw_close.get(sym, {}) or {}).get(i)
        r_ps = pos["entry_px"] - pos["stop_px"]
        R = (mark - pos["entry_px"]) / r_ps if (mark and r_ps) else None
        unreal = (mark - pos["entry_px"]) * pos["lots"] * 100 if mark else None
        rows.append([sym, pos.get("sector"), pos.get("opened"), pos["lots"],
                     pos["entry_px"], pos["stop_px"], mark, R, unreal,
                     (mark or pos["entry_px"]) * pos["lots"] * 100, pos.get("r_idr"),
                     (mark - pos["stop_px"]) / mark if mark else None,
                     pos.get("entry_rule"), (q or {}).get("source", "panel")])
    return cols, rows


def sheet_ledger(events: list) -> tuple:
    """The raw events, unprocessed. Everything else here is derived; this is the source."""
    keys = []
    for e in events:
        for k in e:
            if k != "_line" and k not in keys:
                keys.append(k)
    cols = ["line"] + keys
    rows = [[e.get("_line")] + [e.get(k) for k in keys] for e in events]
    return cols, rows


def sheet_equity(events: list, p: Panel) -> tuple:
    """Daily marked equity from the first fill onward.

    equity(d) = base equity + realised to date + unrealised at that session's close.
    Reconstructed from the ledger rather than sampled live, so it is reproducible and
    does not depend on anything having been running on a given day.
    """
    cols = ["date", "equity_idr", "realised_to_date", "unrealised", "n_open", "gross_pct"]
    if not p.dates:
        return cols, []
    base, realised = 0.0, 0.0
    holdings: dict = {}
    timeline: dict = {}
    for e in events:
        d = e.get("date") or (e.get("ts") or "")[:10]
        timeline.setdefault(d, []).append(e)
    first = min(timeline) if timeline else None
    if not first:
        return cols, []

    rows = []
    for i, d in enumerate(p.dates):
        if d < first:
            continue
        for e in timeline.get(d, []):
            t = e.get("type")
            if t == "equity":
                base = float(e["equity_idr"])
            elif t == "open":
                holdings[e["symbol"]] = {"lots": e["lots"], "entry": e["entry_px"]}
            elif t in ("scale", "close"):
                realised += float(e.get("realised_idr") or 0)
                h = holdings.get(e["symbol"])
                if h:
                    h["lots"] += e["lots"]          # lots is negative on an exit
                    if h["lots"] <= 0:
                        holdings.pop(e["symbol"], None)
        unreal, gross = 0.0, 0.0
        for sym, h in holdings.items():
            px = (p.raw_close.get(sym, {}) or {}).get(i)
            if px:
                unreal += (px - h["entry"]) * h["lots"] * 100
                gross += px * h["lots"] * 100
        eq = base + realised + unreal
        rows.append([d, eq, realised, unreal, len(holdings),
                     gross / eq if eq else None])
    return cols, rows


def sheet_boardfires(p: Panel, events: list, horizons=(3, 5, 10)) -> tuple:
    """Every board candidate the panel can reconstruct, and whether you took it.

    The `taken` join is by (symbol, entry session) with a two-session tolerance: the
    convention is entry at the close AFTER the signal, but a fill recorded a day late
    is still that trade, and treating it as a different one would understate what you
    actually took.
    """
    import build_momentum_board as bmb
    import trade_backtest as tb
    cands = tb.build_candidates(p)
    # build_candidates goes through build_events, which DROPS any event whose forward
    # horizons are unavailable — so it stops ~14 sessions before the panel end. Correct
    # for a backtest and useless here: every position in the book was opened inside that
    # blind spot, so the selection question came back "0 taken, 1,563 skipped" and looked
    # like an answer rather than a gap.
    #
    # The board's own builder scores ONE session and needs no forward return, so the tail
    # is filled from build_momentum_board.build() — the same gates that actually fired,
    # not a re-derivation. Those rows carry null forward returns until the sessions
    # exist, which is honest: the outcome is genuinely not known yet.
    last_i = max((c["i"] for c in cands), default=-1)
    seen_k = {(c["symbol"], c["i"]) for c in cands}
    for j in range(last_i + 1, len(p.dates)):
        try:
            fired, _ex = bmb.build(p, j, {})
        except Exception:                                      # noqa: BLE001
            continue
        for row in bmb.collapse(fired):
            if (row["symbol"], j) in seen_k:
                continue
            seen_k.add((row["symbol"], j))
            cands.append({"symbol": row["symbol"], "i": j,
                          "adtv_pct": row.get("adtv_pct_total"),
                          "n_brokers": len(row.get("brokers") or []),
                          "f": {"rvol5": row.get("rvol"), "rsi": row.get("rsi"),
                                "dd60": row.get("dd60"), "clv": row.get("clv"),
                                "trend": row.get("trend")}})
    cands.sort(key=lambda c: (c["i"], c["symbol"]))
    taken = {}
    for e in events:
        if e.get("type") != "open":
            continue
        d = e.get("date") or (e.get("ts") or "")[:10]
        taken.setdefault(e["symbol"], []).append(d)

    cols = ["signal_date", "symbol", "entry_date", "entry_close", "adtv_pct",
            "n_brokers", "rvol5", "rsi", "dd60", "clv", "trend"] + \
           [f"excess_{k}d" for k in horizons] + ["taken", "taken_date", "outcome_known"]
    rows = []
    for c in cands:
        sym, i = c["symbol"], c["i"]
        ent = i + 1
        if ent >= len(p.dates):
            continue
        f = c.get("f") or {}
        was, tdate = False, ""
        for d in taken.get(sym, []):
            if d in p.didx and abs(p.didx[d] - ent) <= 2:
                was, tdate = True, d
                break
        rows.append([p.dates[i], sym, p.dates[ent],
                     (p.raw_close.get(sym, {}) or {}).get(ent),
                     c.get("adtv_pct"), c.get("n_brokers"),
                     f.get("rvol5"), f.get("rsi"), f.get("dd60"),
                     f.get("clv"), f.get("trend")] +
                    [p.excess_return(sym, i, k) for k in horizons] +
                    [was, tdate, p.excess_return(sym, i, max(horizons)) is not None])
    return cols, rows


def sheet_summary(state: dict, fires: tuple) -> tuple:
    cols = ["group", "key", "n", "mean_R", "median_R", "hit_rate", "total_idr"]
    rows = []

    def add(group, key, trades):
        if not trades:
            return
        R = [t.get("realised_r_total") for t in trades if t.get("realised_r_total") is not None]
        idr = sum(t.get("realised_idr") or 0 for t in trades)
        rows.append([group, key, len(trades),
                     statistics.fmean(R) if R else None,
                     statistics.median(R) if R else None,
                     (sum(1 for x in R if x > 0) / len(R)) if R else None, idr])

    closed = state["closed"]
    add("all", "closed trades", closed)
    for k in ("entry_rule", "sector", "exit_reason"):
        for v in sorted({str(c.get(k)) for c in closed}):
            add(k, v, [c for c in closed if str(c.get(k)) == v])

    # The selection question, which is the reason BoardFires exists.
    fcols, frows = fires
    if frows and "taken" in fcols:
        ti, ei = fcols.index("taken"), fcols.index("excess_5d")
        took = [r[ei] for r in frows if r[ti] and r[ei] is not None]
        skip = [r[ei] for r in frows if not r[ti] and r[ei] is not None]
        for label, xs in (("board fires TAKEN", took), ("board fires SKIPPED", skip)):
            if xs:
                rows.append(["selection", label, len(xs), None, None,
                             sum(1 for x in xs if x > 0) / len(xs),
                             statistics.fmean(xs)])
        # Say what is NOT yet comparable. A selection statistic computed over rows whose
        # outcome is unknown would silently answer a question the data cannot.
        pending = [r for r in frows if r[ti] and r[ei] is None]
        if pending:
            rows.append(["selection", "TAKEN, outcome not yet known", len(pending),
                         None, None, None, None])
    return cols, rows


# ------------------------------------------------------------------------------- write

def write_workbook(out: Path, sheets: dict) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)
    head_font = Font(bold=True, color="FFFFFFFF")
    head_fill = PatternFill("solid", fgColor="FF1F2937")
    for name, (cols, rows) in sheets.items():
        ws = wb.create_sheet(name[:31])
        ws.append(cols)
        for c in ws[1]:
            c.font, c.fill = head_font, head_fill
            c.alignment = Alignment(horizontal="left")
        for r in rows:
            ws.append([v if not isinstance(v, (dict, list)) else json.dumps(v) for v in r])
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for j, col in enumerate(cols, 1):
            width = max(len(str(col)) + 2,
                        *(len(str(r[j - 1])) + 2 for r in rows[:200]) if rows else [10])
            ws.column_dimensions[get_column_letter(j)].width = min(28, max(9, width))
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument("--out", type=str, default=None)
    ap.add_argument("--no-boardfires", action="store_true",
                    help="skip the panel reconstruction (much faster)")
    a = ap.parse_args()

    today = datetime.now(WIB).strftime("%Y-%m-%d")
    out = Path(a.out) if a.out else (pb.BOOK / "exports" / f"book-{today}.xlsx")
    _refuse_public(out)

    events = pb.read_events()
    state = pb.rebuild(events)
    print(f"ledger: {len(events)} events | {len(state['positions'])} open | "
          f"{len(state['closed'])} closed")

    p = Panel()
    if a.no_boardfires:
        p.load_prices()
    else:
        p.load()
        p.load_benchmark()
    print(f"panel: {len(p.dates)} sessions to {p.dates[-1] if p.dates else '—'}")

    marks = {}
    if state["positions"]:
        import live_prices
        marks = live_prices.quotes(sorted(state["positions"]))

    fires = ([], [])
    if not a.no_boardfires:
        print("reconstructing board fires...")
        fires = sheet_boardfires(p, events)
        print(f"  {len(fires[1])} candidates")

    sheets = {
        "Summary": sheet_summary(state, fires),
        "Open": sheet_open(state, p, marks),
        "Trades": sheet_trades(state),
        "Equity": sheet_equity(events, p),
        "Ledger": sheet_ledger(events),
    }
    if not a.no_boardfires:
        sheets["BoardFires"] = fires

    write_workbook(out, sheets)
    size = out.stat().st_size
    print(f"\nwrote {out}  ({size / 1024:.0f} KB)")
    for n, (c, r) in sheets.items():
        print(f"  {n:<12}{len(r):>6} rows x {len(c)} cols")
    return 0


if __name__ == "__main__":
    sys.exit(main())
